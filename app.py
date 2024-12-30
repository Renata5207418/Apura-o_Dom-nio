from pywinauto import Application, findwindows
from flask import Flask, request, jsonify, render_template, send_from_directory
from datetime import datetime
from openpyxl import Workbook
from dotenv import load_dotenv
import time
import openpyxl
import subprocess
import pyautogui
import os
import psutil

# Carregar variáveis do .env
load_dotenv()
dominio_process = None
flask_app = Flask(__name__)

# Caminhos e credenciais
CAMINHO_EXECUTAVEL = os.getenv("CAMINHO_EXECUTAVEL")
UPLOAD_FOLDER = os.getenv("UPLOAD_FOLDER", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

progresso = []


@flask_app.route("/")
def index():
    """Serve a página principal."""
    global progresso
    progresso.clear()    # Reseta o progresso ao abrir a página
    return render_template("index.html")


def adicionar_progresso(mensagem, limpar=False):
    """
    Adiciona uma mensagem à lista de progresso.
    """
    global progresso
    if limpar:
        progresso.clear()
    progresso.append(mensagem)
    print(mensagem)


@flask_app.route("/upload-and-login", methods=["POST"])
def upload_and_login():
    """
        Rota responsável por:
        1. Processar o upload do arquivo Excel com os dados das empresas.
        2. Realizar login no sistema Domínio.
        3. Processar apuração para cada empresa listada no arquivo.
    """
    global progresso
    progresso.clear()
    resultados = []

    try:
        usuario = request.form["usuario"]
        senha = request.form["senha"]
        file = request.files["arquivo_empresas"]

        if not file:
            return jsonify({"error": "Nenhum arquivo enviado"}), 400

        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        progresso.append("Abrindo o sistema Domínio...")
        abrir_dominio(CAMINHO_EXECUTAVEL)
        app_dominio = Application(backend="uia").connect(path=CAMINHO_EXECUTAVEL)
        progresso.append("Realizando login...")
        fazer_login(app_dominio, usuario, senha)

        progresso.append("Carregando dados da planilha...")
        dados = carregar_planilha(file_path)
        progresso.append(f"{len(dados)} empresas carregadas para processamento.")

        for entrada in dados:
            codigo_empresa = entrada["codigo_empresa"]
            mes_ano = entrada["mes_ano"]
            try:
                progresso.append(f"Processando empresa {codigo_empresa} para {mes_ano.strftime('%m/%Y')}...")
                selecionar_empresa(app_dominio, codigo_empresa)
                status = processar_apuracao(app_dominio, mes_ano, codigo_empresa)

                if status == "Aviso":
                    mensagem_aviso = f"Empresa {codigo_empresa} com aviso. Necessário verificar manualmente."
                    progresso.append(mensagem_aviso)
                    resultados.append({"codigo_empresa": codigo_empresa, "status": "Aviso de Apuração (Verificar)"})
                elif status == "Sucesso":
                    mensagem_sucesso = f"Empresa {codigo_empresa} processada com sucesso."
                    progresso.append(mensagem_sucesso)
                    resultados.append({"codigo_empresa": codigo_empresa, "status": "Processada"})

            except Exception as e:
                progresso.append(f"Erro ao processar a empresa {codigo_empresa}: {e}")
                resultados.append({"codigo_empresa": codigo_empresa, "status": "Erro"})
                progresso.append("Reiniciando o sistema para tentar novamente...")
                fechar_dominio()
                abrir_dominio(CAMINHO_EXECUTAVEL)
                app_dominio = Application(backend="uia").connect(path=CAMINHO_EXECUTAVEL)
                fazer_login(app_dominio, usuario, senha)

        progresso.append("Gerando relatório final de apuração...")
        relatorio_path = gerar_relatorio(resultados)

        relatorio_nome = os.path.basename(relatorio_path)

        progresso.append("Fechando o sistema Domínio...")
        fechar_dominio()

        return jsonify(
            {"message": "Login e processamento realizados com sucesso!", "relatorio": relatorio_nome}
        ), 200

    except Exception as e:
        progresso.append(f"Erro: {e}")
        return jsonify({"error": str(e)}), 500


@flask_app.route("/uploads/<filename>")
def download_file(filename):
    """Permite o download do arquivo de relatório gerado"""
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


@flask_app.route("/progresso", methods=["GET"])
def progresso_endpoint():
    """Rota para retornar mensagens de progresso."""
    global progresso
    return jsonify(progresso), 200


def gerar_relatorio(resultados):
    """Gera um Relatório ao final em um arquivo xlsx com o resultado dos processamentos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Apuração"
    ws.append(["Código da Empresa", "Status"])

    for resultado in resultados:
        ws.append([resultado["codigo_empresa"], resultado["status"]])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    relatorio_path = os.path.join(UPLOAD_FOLDER, f"relatorio_apuracao_{timestamp}.xlsx")
    wb.save(relatorio_path)

    return relatorio_path


# Daqui em diante, funções auxiliares que são necessárias para abrir, fechar e interagir com o sistema Domínio.

def abrir_dominio(caminho_executavel):
    """
    Abre o sistema Domínio usando o caminho configurado no .env.
    Aguarda o carregamento antes de prosseguir.
    """
    global dominio_process
    try:
        dominio_process = subprocess.Popen([caminho_executavel, '/escrita'])
        time.sleep(10)
    except Exception as e:
        raise Exception(f"Erro ao tentar abrir o sistema: {e}")


def fazer_login(app, usuario, senha):
    """
    Realiza login no sistema Domínio com o usuário e senha fornecidos.
    Localiza os campos de login na interface gráfica usando pywinauto.
    """

    try:
        login_window = app.window(title_re="Conectando.*")
        login_window.wait('visible', timeout=30)

        usuario_field = login_window.child_window(auto_id="1005")
        usuario_field.click_input(double=True)
        usuario_field.type_keys("{DEL}")
        usuario_field.type_keys(usuario)

        login_window.child_window(auto_id="1007").type_keys(senha)
        login_window.child_window(auto_id="1003", title="OK").click()
        time.sleep(15)
        progresso.append("Login realizado com sucesso.")
    except Exception as e:
        raise Exception(f"Erro ao realizar o login: {e}")


def selecionar_empresa(app, codigo_empresa):
    """
    Seleciona uma empresa pelo código fornecido.
    Usa coordenadas na interface gráfica para preencher os campos.
    """

    try:
        main_window = app.window(title_re=".*Domínio.*")
        main_window.set_focus()
        time.sleep(2)

        main_window.type_keys("{F8}")
        time.sleep(3)

        codigo_radiobutton = main_window.child_window(auto_id="1", control_type="RadioButton")
        codigo_radiobutton.click_input()
        time.sleep(1)

        codigo_edit = main_window.child_window(auto_id="1002", control_type="Edit")
        codigo_edit.set_focus()
        codigo_edit.type_keys(codigo_empresa)
        time.sleep(1)

        pyautogui.press("enter")
        time.sleep(2)

    except Exception as e:
        raise Exception(f"Erro ao selecionar a empresa {codigo_empresa}: {e}")


def processar_apuracao(app, mes_ano, codigo_empresa):
    """
    Processa a apuração da empresa no sistema Domínio para o mês/ano informado.
    Detecta janelas de aviso e finaliza automaticamente.
    """
    try:
        print(f"Iniciando o processamento de apuração para {mes_ano.strftime('%m/%Y')}...")

        aviso_apuracao_detectado = False

        main_window = app.window(title_re=".*Domínio.*")
        main_window.set_focus()
        time.sleep(2)

        main_window.type_keys("%M")
        main_window.type_keys("R")
        print("Comando para abrir a janela de Apuração enviado.")
        time.sleep(3)

        data_final_edit = main_window.child_window(auto_id="1009", control_type="Edit")
        data_final_edit.set_focus()
        data_final_edit.type_keys(mes_ano.strftime("%m%Y"))
        time.sleep(2)

        main_window.type_keys("%P")
        print("Processamento iniciado. Aguardando finalização...")

        try:
            atencao_windows = findwindows.find_windows(title="Atenção", class_name="#32770")
            if atencao_windows:
                atencao_window = app.window(handle=atencao_windows[0])
                atencao_window.set_focus()
                print(f"Janela 'Atenção' detectada para empresa {codigo_empresa}. Confirmando para continuar.")
                pyautogui.press("S")
                time.sleep(2)
        except findwindows.ElementNotFoundError:
            pass

        timeout = 100
        start_time = time.time()
        apuracao_window = None

        while time.time() - start_time < timeout:
            try:
                aviso_windows = findwindows.find_windows(title="Avisos da Apuração", class_name="FNWNS3190")
                if aviso_windows:
                    aviso_apuracao_window = app.window(handle=aviso_windows[0])
                    aviso_apuracao_window.set_focus()
                    mensagem_aviso = (f"Aviso: Janela 'Avisos da Apuração' detectada para empresa {codigo_empresa}."
                                      f" Necessário verificar manualmente.")
                    progresso.append(mensagem_aviso)
                    print(mensagem_aviso)
                    aviso_apuracao_window.close()
                    aviso_apuracao_detectado = True
                    time.sleep(2)

                windows = findwindows.find_windows(title="Fim da apuração!", class_name="#32770")
                if windows:
                    apuracao_window = app.window(handle=windows[0])
                    apuracao_window.set_focus()
                    print("Janela 'Fim da apuração!' detectada.")
                    break
            except findwindows.ElementNotFoundError:
                pass

            time.sleep(1)

        if not apuracao_window:
            raise TimeoutError("A janela 'Fim da apuração!' não apareceu no tempo esperado.")

        apuracao_window.set_focus()
        pyautogui.press("N")
        print("Mensagem de fim da apuração fechada.")

        main_window.set_focus()
        pyautogui.press("ESC")
        print("Janela de Apuração fechada.")

        if aviso_apuracao_detectado:
            return "Aviso"
        else:
            return "Sucesso"

    except Exception as e:
        print(f"Erro ao processar a apuração para {mes_ano.strftime('%m/%Y')}: {e}")
        raise e


def fechar_dominio():
    """
    Fecha o sistema Domínio e verifica se o processo ainda está ativo.
    Finaliza o processo, se necessário.
    """

    global dominio_process
    try:
        if dominio_process is not None:
            pyautogui.hotkey('alt', 'f4')
            time.sleep(5)

            if dominio_process.poll() is None:
                for proc in psutil.process_iter(['pid', 'name']):
                    if proc.name().lower() == "contabil.exe":
                        proc.terminate()
            dominio_process = None
    except Exception as e:
        progresso.append(f"Erro ao fechar o sistema: {e}")


def carregar_planilha(caminho_planilha):
    """Carrega dados da planilha e retorna como lista de dicionários."""
    try:
        wb = openpyxl.load_workbook(caminho_planilha)
        sheet = wb.active
        dados = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_empresa, mes_ano = row

            if codigo_empresa is not None and mes_ano is not None:
                dados.append({"codigo_empresa": str(codigo_empresa), "mes_ano": mes_ano})

        return dados
    except Exception as e:
        raise Exception(f"Erro ao carregar a planilha: {e}")


if __name__ == "__main__":
    flask_app.run(debug=True)
